# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
import openpyxl.worksheet
import unicodedata
from string import ascii_letters
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image
import time
from datetime import datetime,timedelta,date
import calendar
import locale
from odoo.exceptions import UserError, ValidationError


def informe_credito_cobranza(ruta,lista,clave):

    workbook = openpyxl.load_workbook(ruta)

    sheet = workbook.active

    if clave=='orden_compra':

        sheet = workbook['OrdenCompra']
        listaSheet = list(filter(lambda x: (x['hoja']==1), lista)) 
        for campo in listaSheet:
            cell = sheet.cell(row=campo['fila'], column=campo['columna'])
            try:
                cell.value = campo['valor'] or ''
            except:
                raise ValidationError("""El valor {0} en la fila {1} columna {2}  hoja {3} se encuentra mal configurado en la plantilla""".format( campo['valor'], campo['fila'],campo['columna'],campo['hoja']))
        
    elif clave=='orden_salida':
        sheet = workbook['OrdenSalida']
        listaSheet = list(filter(lambda x: (x['hoja']==1), lista)) 
        for campo in listaSheet:
            cell = sheet.cell(row=campo['fila'], column=campo['columna'])
            try:
                cell.value = campo['valor'] or ''
            except:
                raise ValidationError("""El valor {0} en la fila {1} columna {2}  hoja {3} se encuentra mal configurado en la plantilla""".format( campo['valor'], campo['fila'],campo['columna'],campo['hoja']))

    workbook.save(ruta)
    return workbook
