# -*- coding: utf-8 -*-s

from odoo import api, fields, models, tools
from datetime import date, timedelta,datetime
from dateutil.relativedelta import relativedelta
import xlsxwriter
from io import BytesIO
import base64
from odoo.exceptions import AccessError, UserError, ValidationError
#from . import l10n_ec_check_printing.amount_to_text_es
from . import amount_to_text_es
from datetime import datetime
import calendar
import datetime as tiempo
import itertools
from . import contrato_reserva_documento
import shutil

import os
import re
import json
import base64
import logging
import mimetypes
import odoo.tools
import hashlib
from odoo import api, fields, models, tools, SUPERUSER_ID
from datetime import datetime,timedelta,date
import time
from odoo import _
from odoo.exceptions import ValidationError, except_orm
from dateutil.relativedelta import *
from . import orden_compra_salida_documento
import subprocess
from subprocess import getoutput
import os
import io
import base64
from base64 import urlsafe_b64decode
import openpyxl
from openpyxl import Workbook
import openpyxl.worksheet
import unicodedata
from string import ascii_letters
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image
import time
from datetime import datetime,timedelta,date
import calendar
import locale
from odoo.exceptions import UserError, ValidationError

import shutil



class EntrevistaCredito(models.TransientModel):
    _name = "entrevista.credito"

    entrega_vehiculo_id =  fields.Many2one('entrega.vehiculo',string='Entrega Vehiculo',)
    clave =  fields.Char( default="")
    archivo_xls1 = fields.Binary('Archivo excel')


    def crear_formato_entrevista(self,):
        #Instancia la plantilla
        garante=False
        obj_plantilla=self.env['plantillas.dinamicas.informes'].search([('identificador_clave','=',self.clave)],limit=1)
            
        if obj_plantilla:
            shutil.copy2(obj_plantilla.directorio,obj_plantilla.directorio_out)
            campos=obj_plantilla.campos_ids.filtered(lambda l: len(l.child_ids)==0)
            lista_campos=[]
            for campo in campos:
                resultado=self.entrega_vehiculo_id.mapped(campo.name)
                dct={}
                if len(resultado)>0:
                    dct['valor']=resultado[0]
                else:
                    dct['valor']=''

                dct['fila']=campo.fila
                dct['columna']=campo.columna
                dct['hoja']=campo.hoja_excel
                lista_campos.append(dct)
            

            dct={}
            if self.entrega_vehiculo_id.contrato_id.grupo:
                grupo_contrato=self.entrega_vehiculo_id.contrato_id.grupo.name+"/"+self.entrega_vehiculo_id.contrato_id.secuencia
                dct['fila']=1
                dct['columna']=6
                dct['hoja']=1
                dct['valor']=grupo_contrato
                lista_campos.append(dct)
            dct={}
            dct['fila']=2
            dct['columna']=6
            dct['hoja']=1
            dct['valor']=self.entrega_vehiculo_id.contrato_id.fecha_adjudicado
            lista_campos.append(dct)

            if self.entrega_vehiculo_id:
                if self.entrega_vehiculo_id.estadoVehiculo=="USADO":
                    inicia=62
                elif self.entrega_vehiculo_id.estadoVehiculo=="NUEVO":
                    inicia=66
                cont=1
                for detalle in self.entrega_vehiculo_id.detalleVehiculoIds:
                    if cont==1:
                        dct={}
                        dct['hoja']=1
                        dct['fila']=inicia
                        dct['columna']=2
                        dct['valor']=detalle.marca
                        lista_campos.append(dct)
                        dct={}
                        dct['hoja']=1
                        dct['fila']=inicia+1
                        dct['columna']=2
                        dct['valor']=detalle.modelo
                        lista_campos.append(dct)
                        dct={}
                        dct['hoja']=1
                        dct['fila']=inicia+2
                        dct['columna']=2
                        dct['valor']=detalle.anio
                        lista_campos.append(dct)
                        dct={}
                        dct['hoja']=1
                        dct['fila']=inicia+3
                        dct['columna']=2
                        dct['valor']=detalle.colores
                        lista_campos.append(dct)
                        break
                for necesidad in self.entrega_vehiculo_id.necesidadIds:
                    if necesidad.necesidad_id.celda_excel in [62,63,64,65,66,67,68,69]:
                        dct={}
                        dct['hoja']=1
                        dct['fila']=necesidad.necesidad_id.celda_excel
                        dct['columna']=2
                        dct['valor']="No"
                        if necesidad.adquirirBien:
                            dct['valor']="SI"
                        lista_campos.append(dct)

                if self.clave=="entrevista_adjudicado":
                    for items in self.entrega_vehiculo_id.montoAhorroInversiones:
                        if not items.garante:
                            if items.patrimonio_id.celda_excel in [28,29,30,31,32,33,34]:
                                dct={}
                                dct['hoja']=1
                                dct['fila']=items.patrimonio_id.celda_excel
                                dct['columna']=2
                                dct['valor']=items.poseeBien
                                lista_campos.append(dct)
                                dct={}
                                dct['hoja']=1
                                dct['fila']=items.patrimonio_id.celda_excel
                                dct['columna']=3
                                dct['valor']=items.descripcion
                                lista_campos.append(dct)
                                dct={}
                                dct['hoja']=1
                                dct['fila']=items.patrimonio_id.celda_excel
                                dct['columna']=5
                                dct['valor']=items.valor
                                lista_campos.append(dct)
                    
                    for ingreso in self.entrega_vehiculo_id.ingresosIds:
                        if not ingreso.garante:
                            if ingreso.ingresos_id.celda_excel in [37,38]:
                                dct={}
                                dct['hoja']=1
                                dct['fila']=ingreso.ingresos_id.celda_excel
                                dct['columna']=5
                                dct['valor']=ingreso.titular
                                lista_campos.append(dct)
                                dct={}
                                dct['hoja']=1
                                dct['fila']=ingreso.ingresos_id.celda_excel
                                dct['columna']=6
                                dct['valor']=ingreso.conyuge
                                lista_campos.append(dct)

                    for egreso in self.entrega_vehiculo_id.egresosIds:
                        if not egreso.garante:
                            if egreso.egresos_id.celda_excel in [41,42,43,44,45,46,47,48,49]:
                                dct={}
                                dct['hoja']=1
                                dct['fila']=egreso.egresos_id.celda_excel
                                dct['columna']=5
                                dct['valor']=egreso.titular
                                lista_campos.append(dct)
                                dct={}
                                dct['hoja']=1
                                dct['fila']=egreso.egresos_id.celda_excel
                                dct['columna']=5
                                dct['valor']=egreso.conyuge
                                lista_campos.append(dct)
                    i=0
                    for familia in self.entrega_vehiculo_id.referencias_familiares_ids:
                        if i<2:
                            if not familia.garante:
                                dct={}
                                dct['hoja']=1
                                dct['fila']=55+i
                                dct['columna']=1
                                dct['valor']=familia.nombre
                                lista_campos.append(dct)
                                dct={}
                                dct['hoja']=1
                                dct['fila']=55+i
                                dct['columna']=2
                                dct['valor']=familia.cedula
                                lista_campos.append(dct)
                                dct={}
                                dct['hoja']=1
                                dct['fila']=55+i
                                dct['columna']=3
                                dct['valor']=familia.parentezco
                                lista_campos.append(dct)
                                dct={}
                                dct['hoja']=1
                                dct['fila']=55+i
                                dct['columna']=4
                                dct['valor']=familia.direccion
                                lista_campos.append(dct)
                                dct={}
                                dct['hoja']=1
                                dct['fila']=55+i
                                dct['columna']=5
                                dct['valor']=familia.telefono
                                lista_campos.append(dct)
                                i+=1

                    j=0
                    for bancaria in self.entrega_vehiculo_id.referencias_bancarias_ids:
                        if j<2:
                            if not bancaria.garante:
                                dct={}
                                dct['hoja']=1
                                dct['fila']=59+j
                                dct['columna']=1
                                dct['valor']=bancaria.banco
                                lista_campos.append(dct)
                                dct={}
                                dct['hoja']=1
                                dct['fila']=59+j
                                dct['columna']=2
                                dct['valor']=bancaria.tipo_cuenta
                                lista_campos.append(dct)
                                dct={}
                                dct['hoja']=1
                                dct['fila']=59+j
                                dct['columna']=4
                                dct['valor']=bancaria.numero_cuenta
                                lista_campos.append(dct)
                                j+=1
                else:
                    for items in self.entrega_vehiculo_id.montoAhorroInversionesGarante:
                        if items.garante:
                            if items.patrimonio_id.celda_excel in [28,29,30,31,32,33,34]:
                                dct={}
                                dct['hoja']=1
                                dct['fila']=items.patrimonio_id.celda_excel
                                dct['columna']=2
                                dct['valor']=items.poseeBien
                                lista_campos.append(dct)
                                dct={}
                                dct['hoja']=1
                                dct['fila']=items.patrimonio_id.celda_excel
                                dct['columna']=3
                                dct['valor']=items.descripcion
                                lista_campos.append(dct)
                                dct={}
                                dct['hoja']=1
                                dct['fila']=items.patrimonio_id.celda_excel
                                dct['columna']=5
                                dct['valor']=items.valor
                                lista_campos.append(dct)
                
                    for ingreso in self.entrega_vehiculo_id.ingresosIdsGarante:
                        if ingreso.garante:
                            if ingreso.ingresos_id.celda_excel in [37,38]:
                                dct={}
                                dct['hoja']=1
                                dct['fila']=ingreso.ingresos_id.celda_excel
                                dct['columna']=5
                                dct['valor']=ingreso.titular
                                lista_campos.append(dct)
                                dct={}
                                dct['hoja']=1
                                dct['fila']=ingreso.ingresos_id.celda_excel
                                dct['columna']=6
                                dct['valor']=ingreso.conyuge
                                lista_campos.append(dct)
                    for egreso in self.entrega_vehiculo_id.egresosIdsGarante:
                        if egreso.garante:
                            if egreso.egresos_id.celda_excel in [41,42,43,44,45,46,47,48,49]:
                                dct={}
                                dct['hoja']=1
                                dct['fila']=egreso.egresos_id.celda_excel
                                dct['columna']=5
                                dct['valor']=egreso.titular
                                lista_campos.append(dct)
                                dct={}
                                dct['hoja']=1
                                dct['fila']=egreso.egresos_id.celda_excel
                                dct['columna']=6
                                dct['valor']=egreso.conyuge
                                lista_campos.append(dct)
                    i=0
                    for familia in self.entrega_vehiculo_id.referencias_familiares_ids_garante:
                        if i<2:
                            if familia.garante:
                                dct={}
                                dct['hoja']=1
                                dct['fila']=55+i
                                dct['columna']=1
                                dct['valor']=familia.nombre
                                lista_campos.append(dct)
                                dct={}
                                dct['hoja']=1
                                dct['fila']=55+i
                                dct['columna']=2
                                dct['valor']=familia.cedula
                                lista_campos.append(dct)
                                dct={}
                                dct['hoja']=1
                                dct['fila']=55+i
                                dct['columna']=3
                                dct['valor']=familia.parentezco
                                lista_campos.append(dct)
                                dct={}
                                dct['hoja']=1
                                dct['fila']=55+i
                                dct['columna']=4
                                dct['valor']=familia.direccion
                                lista_campos.append(dct)
                                dct={}
                                dct['hoja']=11
                                dct['fila']=55+i
                                dct['columna']=5
                                dct['valor']=familia.telefono
                                lista_campos.append(dct)
                                i+=1
                    j=0
                    for bancaria in self.entrega_vehiculo_id.referencias_bancarias_ids_garante:
                        if j<2:
                            if bancaria.garante:
                                dct={}
                                dct['hoja']=1
                                dct['fila']=59+j
                                dct['columna']=1
                                dct['valor']=bancaria.banco
                                lista_campos.append(dct)
                                dct={}
                                dct['hoja']=1
                                dct['fila']=59+j
                                dct['columna']=2
                                dct['valor']=bancaria.tipo_cuenta
                                lista_campos.append(dct)
                                dct={}
                                dct['hoja']=1
                                dct['fila']=59+j
                                dct['columna']=4
                                dct['valor']=bancaria.numero_cuenta
                                lista_campos.append(dct)
                                j+=1
            fp =io.BytesIO()
            workbook=self.entrevista_credito(obj_plantilla.directorio_out,lista_campos)
            workbook.save(fp)
            with open(obj_plantilla.directorio_out, "rb") as f:
                data = f.read()
                file=bytes(base64.b64encode(data))

        nombre_doc="Entrevista.xlsx"

        obj_attch=self.env['ir.attachment'].create({
                                                    'name':nombre_doc,
                                                    'datas':file,
                                                    'type':'binary', 
                                                    'store_fname':nombre_doc
                                                    })

        
        direccion_xls_libro=self.env['ir.attachment']._get_path(obj_attch.datas,obj_attch.checksum)[1]
        nombre_bin=obj_attch.checksum
        nombre_archivo=obj_attch.name
        os.chdir(direccion_xls_libro.rstrip(nombre_bin))
        print(os.chdir(direccion_xls_libro.rstrip(nombre_bin)))
        os.rename(nombre_bin,nombre_archivo)
        subprocess.getoutput("""libreoffice --headless --convert-to pdf *.xlsx""") 
        try:
            with open(direccion_xls_libro.rstrip(nombre_bin)+nombre_archivo.split('.')[0]+'.pdf', "rb") as f:
                data = f.read()
                file=bytes(base64.b64encode(data))
        except:
            raise ValidationError(_('No existen datos para generar informe'))
        obj_attch.unlink()
        obj_attch=self.env['ir.attachment'].create({
                                                    'name':nombre_archivo.split('.')[0]+'.pdf',
                                                    'datas':file,
                                                    'type':'binary', 
                                                    'store_fname':nombre_archivo.split('.')[0]+'.pdf'
                                                    })
        self.archivo_xls1=file
        url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
        url += "/web/content/%s?download=true"%(obj_attch.id) 
        return{
            "type": "ir.actions.act_url",
            "url": url,
            "target": "new",
        }

    def entrevista_credito(self,ruta,lista):
        workbook = openpyxl.load_workbook(ruta)
        sheet = workbook.active
        sheet = workbook['Entrevista']
        listaSheet = list(filter(lambda x: (x['hoja']==1), lista)) 
        for campo in listaSheet:
            cell = sheet.cell(row=campo['fila'], column=campo['columna'])
            try:
                cell.value = campo['valor'] or ''
            except:
                raise ValidationError("""El valor {0} en la fila {1} columna {2}  hoja {3} se encuentra mal configurado en la plantilla""".format( campo['valor'], campo['fila'],campo['columna'],campo['hoja']))
        workbook.save(ruta)
        return workbook