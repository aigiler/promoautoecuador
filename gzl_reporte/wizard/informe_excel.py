# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
import openpyxl.worksheet
import unicodedata
from string import ascii_letters
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image
import time
from datetime import datetime,timedelta,date
import calendar
import locale
from odoo.exceptions import UserError, ValidationError



# def crear_wb():
#     wb = Workbook(write_only=False, iso_dates=False)
#     return wb


# def unicodeText(text):
#     try:
#         text = unicodedata.unicode(text, 'utf-8')
#     except TypeError:
#         return text

# def crea_hoja(wb, title, flag):
#     if(flag == 0):
#         sheet = wb.active
#         sheet.sheet_properties.pageSetUpPr.fitToPage = True

#         sheet.page_setup.fitToWidht = False
#     if(flag == 1):
#         sheet = wb.create_sheet()
#         sheet.sheet_properties.pageSetUpPr.fitToPage = True

#         sheet.page_setup.fitToWidht = False
#     sheet.title = title
#     return sheet

# # Ajustar tamanios de celdas
# def ajustar_hoja(sheet, flag, celda, value):
#     if (flag == 0):
#         sheet.column_dimensions[celda].width = value
#     if (flag == 1):
#         sheet.row_dimensions[int(celda)].height = value




def informe_credito_cobranza(ruta,lista,clave):

    workbook = openpyxl.load_workbook(ruta)

    sheet = workbook.active

    if clave=='orden_compra':

        sheet = workbook['OrdenCompra']
        listaSheet = list(filter(lambda x: (x['hoja']==1), lista)) 
        for campo in listaSheet:
            cell = sheet.cell(row=campo['fila'], column=campo['columna'])
            try:
                cell.value = campo['valor'] or ''
            except:
                raise ValidationError("""El valor {0} en la fila {1} columna {2}  hoja {3} se encuentra mal configurado en la plantilla""".format( campo['valor'], campo['fila'],campo['columna'],campo['hoja']))
        
    elif clave=='orden_salida':
        sheet = workbook['OrdenSalida']
        listaSheet = list(filter(lambda x: (x['hoja']==1), lista)) 
        for campo in listaSheet:
            cell = sheet.cell(row=campo['fila'], column=campo['columna'])
            try:
                cell.value = campo['valor'] or ''
            except:
                raise ValidationError("""El valor {0} en la fila {1} columna {2}  hoja {3} se encuentra mal configurado en la plantilla""".format( campo['valor'], campo['fila'],campo['columna'],campo['hoja']))

    workbook.save(ruta)
    return workbook





# def llenar_tabla_excel(lista,sheet,filaInicioBusqueda,filafinBusqueda,ColumnaEje):

#     for linea in lista:
#         nombre_patrimonio=linea['nombre']


#         fila=capturar_fila_de_valor_a_buscar_en_hoja_calculo(sheet,filaInicioBusqueda,filafinBusqueda,ColumnaEje,nombre_patrimonio)
#         if fila:
#             for campo in linea['campos']:
#                 cell = sheet.cell(row=fila, column=campo['columna'])
#                 print(fila,campo['columna'])
#                 cell.value = campo['valor']




# def capturar_fila_de_valor_a_buscar_en_hoja_calculo(sheet,fila_ini,fila_fin,columna_eje,nombre_buscar):


#     for fila in range(fila_ini,fila_fin+1):
#         valor = sheet.cell(row=fila, column=columna_eje)

#         print(valor,nombre_buscar)
#         if valor.value==nombre_buscar:
#             return fila
#     return False



# def Todalatabla(sheet, col, colfin, fil, filfin, styleleft, styletop, styleright, stylebottom):

#     colfin=colfin+1
#     filfin=filfin+2

#     border_cell = Border(left=Side(style=styleleft), top=Side(style=styletop), right=Side(style=styleright), bottom=Side(style=stylebottom))
#     for i in range(fil, filfin-1):
#         for j in range(col, colfin):
#             sheet.cell(row=i, column=j).border = border_cell




# def border_cell(sheet, fil, col, styleleft, styletop, styleright, stylebottom):
#     border_cell = Border(left=Side(style=styleleft), top=Side(style=styletop), right=Side(style=styleright), bottom=Side(style=stylebottom))
#     sheet.cell(row=fil, column=col).border = border_cell
