# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
import openpyxl.worksheet
import unicodedata
from string import ascii_letters
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image
import time
from datetime import datetime,timedelta,date
import calendar
import locale
from odoo.exceptions import UserError, ValidationError



def crear_wb():
    wb = Workbook(write_only=False, iso_dates=False)
    return wb


def unicodeText(text):
    try:
        text = unicodedata.unicode(text, 'utf-8')
    except TypeError:
        return text

def crea_hoja(wb, title, flag):
    if(flag == 0):
        sheet = wb.active
        sheet.sheet_properties.pageSetUpPr.fitToPage = True

        sheet.page_setup.fitToWidht = False
    if(flag == 1):
        sheet = wb.create_sheet()
        sheet.sheet_properties.pageSetUpPr.fitToPage = True

        sheet.page_setup.fitToWidht = False
    sheet.title = title
    return sheet

# Ajustar tamanios de celdas
def ajustar_hoja(sheet, flag, celda, value):
    if (flag == 0):
        sheet.column_dimensions[celda].width = value
    if (flag == 1):
        sheet.row_dimensions[int(celda)].height = value




def generar_hoja_huta(ruta,lista_campos,lista_documentos_postventa, lista_documentos_legal,devolucion_id):

    workbook = openpyxl.load_workbook(ruta)

    sheet = workbook.active

    sheet = workbook['POSVENTA']
    listaSheet1 = list(filter(lambda x: (x['hoja']==1), lista)) 


    for campo in listaSheet1:

        cell = sheet.cell(row=campo['fila'], column=campo['columna'])
        try:
            cell.value = campo['valor'] or ''
        except:
            raise ValidationError("""El valor {0} en la fila {1} columna {2} hoja {3} se encuentra mal configurado en la plantilla""".format( campo['valor'], campo['fila'],campo['columna'],campo['hoja']))

    if devolucion_id.alerta=='CLIENTE':
        sheet['B14'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.alerta=='ABOGADO':
        sheet['C14'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.alerta=='CONSEJO DE LA JUDICATURA':
        sheet['D14'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.alerta=='DEFENSORIA':
        sheet['E14'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.alerta=='FISCALIA':
        sheet['F14'].fill = PatternFill(fill_type="solid", fgColor='009ee5')

    




    if devolucion_id.tipo_devolucion=='DEVOLUCION DE VALORES SIN FIRMAS':
        sheet['C17'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        sheet['C18'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        sheet['C19'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        if devolucion_id.causa_sin_firma_reserva=="NO INTERESADO EN EL CONTRATO":
            sheet['E17'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        elif devolucion_id.causa_sin_firma_reserva=="NO DISPONE DEL DINERO":
            sheet['E18'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        elif devolucion_id.causa_sin_firma_reserva=="OTRO":
            sheet['E19'].fill = PatternFill(fill_type="solid", fgColor='009ee5')




    elif devolucion_id.tipo_devolucion=='DEVOLUCION DE RESERVA':
        sheet['C20'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        sheet['C21'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        sheet['C22'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        sheet['C23'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        if devolucion_id.causa_sin_firma_reserva=="NO INTERESADO EN EL CONTRATO":
            sheet['E20'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        elif devolucion_id.causa_sin_firma_reserva=="NO DISPONE DEL DINERO":
            sheet['E21'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        elif devolucion_id.causa_sin_firma_reserva=="OTRO":
            sheet['E22'].fill = PatternFill(fill_type="solid", fgColor='009ee5')


    elif devolucion_id.tipo_devolucion=='DEVOLUCION DE LICITACION':
        sheet['C24'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        sheet['C25'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        if devolucion_id.causas_licitacion=='NO DESEA NINGUN VEHICULO OFRECIDO':
            sheet['E4'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        elif devolucion_id.causas_licitacion=='NO CUMPLE CON PERFIL DE CREDITO':
            sheet['E25'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        elif devolucion_id.causas_licitacion=='NO CUMPLE CON POLIZA DE SEGURO':
            sheet['E26'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        elif devolucion_id.causas_licitacion=='INSATISFECHO CON EL PROCESO':
            sheet['E27'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        elif devolucion_id.causas_licitacion=='CLIENTE NO ADJUDICADO':
            sheet['E28'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        elif devolucion_id.causas_licitacion=='LICITACION INCOMPLETA':
            sheet['E29'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        elif devolucion_id.causas_licitacion=='OTRO':
            sheet['E30'].fill = PatternFill(fill_type="solid", fgColor='009ee5')




    elif devolucion_id.tipo_devolucion=='DEVOLUCION POR DESISTIMIENTO DEL CONTRATO':
        sheet['C31'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        sheet['C32'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        sheet['C33'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        sheet['C34'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        sheet['C35'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        sheet['C36'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        if devolucion_id.causas_desistimiento=='MALA VENTA':
            sheet['E31'].fill = PatternFill(fill_type="solid", fgColor='009ee5')

        elif devolucion_id.causas_desistimiento=='ENFERMEDAD':
            sheet['E32'].fill = PatternFill(fill_type="solid", fgColor='009ee5')

        elif devolucion_id.causas_desistimiento=='MUERTE':
            sheet['E33'].fill = PatternFill(fill_type="solid", fgColor='009ee5')

        elif devolucion_id.causas_desistimiento=='DESASTRE NATURAL':
            sheet['E34'].fill = PatternFill(fill_type="solid", fgColor='009ee5')

        elif devolucion_id.causas_desistimiento=='NO DISPONE DE LOS RECURSOS':
            sheet['E35'].fill = PatternFill(fill_type="solid", fgColor='009ee5')

        elif devolucion_id.causas_desistimiento=='OTRO':
            sheet['E36'].fill = PatternFill(fill_type="solid", fgColor='009ee5')



    elif devolucion_id.tipo_devolucion=='DEVOLUCION POR CALIDAD DE VENTA':
        sheet['C37'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        sheet['C38'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        sheet['C39'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        sheet['C40'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        if devolucion_id.causas_calidad_venta=='MALA VENTA':
            sheet['E37'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        if devolucion_id.causas_calidad_venta=='NO ESTA DE ACUERDO CON EL PROCESO':
            sheet['E38'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        if devolucion_id.causas_calidad_venta=='ENFERMEDAD':
            sheet['E39'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        if devolucion_id.causas_calidad_venta=='MUERTE':
            sheet['E40'].fill = PatternFill(fill_type="solid", fgColor='009ee5')



    sheet = workbook['LEGAL']
    listaSheet2 = list(filter(lambda x: (x['hoja']==2), lista)) 
    ###########Llenar segundo sheet
    for campo in listaSheet2:
        cell = sheet.cell(row=campo['fila'], column=campo['columna'])
        try:
            cell.value = campo['valor'] or ''
        except:
            raise ValidationError("""El valor {0} en la fila {1} columna {2}  hoja {3} se encuentra mal configurado en la plantilla""".format( campo['valor'], campo['fila'],campo['columna'],campo['hoja']))


    if devolucion_id.alerta=='CLIENTE':
        sheet['B9'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.alerta=='ABOGADO':
        sheet['C9'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.alerta=='CONSEJO DE LA JUDICATURA':
        sheet['D9'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.alerta=='DEFENSORIA':
        sheet['E9'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.alerta=='FISCALIA':
        sheet['F9'].fill = PatternFill(fill_type="solid", fgColor='009ee5')



    if devolucion_id.calidad_venta=='MALA VENTA':
        sheet['B10'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.calidad_venta=='NO ESTA DE ACUERDO CON EL PROCESO':
        sheet['C10'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.calidad_venta=='MUERTE':
        sheet['D10'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.calidad_venta=='ENFERMEDAD':
        sheet['E10'].fill = PatternFill(fill_type="solid", fgColor='009ee5')

    if devolucion_id.tipo_accion=='CLIENTE':
        sheet['C13'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        sheet['C14'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        sheet['C15'].fill = PatternFill(fill_type="solid", fgColor='009ee5')

    elif devolucion_id.tipo_accion=='ABOGADO':
        sheet['C17'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        sheet['C18'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        sheet['C19'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        sheet['C20'].fill = PatternFill(fill_type="solid", fgColor='009ee5')

    elif devolucion_id.tipo_accion=='CONSEJO':
        sheet['C21'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        sheet['C22'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        sheet['C23'].fill = PatternFill(fill_type="solid", fgColor='009ee5')

    elif devolucion_id.tipo_accion=='DEFENSORIA':
        sheet['C25'].fill = PatternFill(fill_type="solid", fgColor='009ee5')

    elif devolucion_id.tipo_accion=='FISCALIA':
        sheet['C29'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
        sheet['C30'].fill = PatternFill(fill_type="solid", fgColor='009ee5')

    elif devolucion_id.tipo_accion=='CAMARA DE COMERCIO':
        sheet['C33'].fill = PatternFill(fill_type="solid", fgColor='009ee5')



    sheet = workbook['ADJUDICACIONES']
    listaSheet3 = list(filter(lambda x: (x['hoja']==3), lista)) 
    for campo in listaSheet3:
        cell = sheet.cell(row=campo['fila'], column=campo['columna'])
        try:
            cell.value = campo['valor'] or ''
        except:
            raise ValidationError("""El valor {0} en la fila {1} columna {2}  hoja {3} se encuentra mal configurado en la plantilla""".format( campo['valor'], campo['fila'],campo['columna'],campo['hoja']))

    if devolucion_id.alerta=='CLIENTE':
        sheet['B8'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.alerta=='ABOGADO':
        sheet['C8'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.alerta=='CONSEJO DE LA JUDICATURA':
        sheet['D8'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.alerta=='DEFENSORIA':
        sheet['E8'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.alerta=='FISCALIA':
        sheet['F8'].fill = PatternFill(fill_type="solid", fgColor='009ee5')



    if devolucion_id.calidad_venta=='MALA VENTA':
        sheet['B9'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.calidad_venta=='NO ESTA DE ACUERDO CON EL PROCESO':
        sheet['C9'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.calidad_venta=='MUERTE':
        sheet['D9'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.calidad_venta=='ENFERMEDAD':
        sheet['E9'].fill = PatternFill(fill_type="solid", fgColor='009ee5')


    sheet = workbook['CONTABILIDAD']
    listaSheet4 = list(filter(lambda x: (x['hoja']==4), lista)) 
    for campo in listaSheet4:
        cell = sheet.cell(row=campo['fila'], column=campo['columna'])
        try:
            cell.value = campo['valor'] or ''
        except:
            raise ValidationError("""El valor {0} en la fila {1} columna {2}  hoja {3} se encuentra mal configurado en la plantilla""".format( campo['valor'], campo['fila'],campo['columna'],campo['hoja']))


    if devolucion_id.alerta=='CLIENTE':
        sheet['B8'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.alerta=='ABOGADO':
        sheet['C8'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.alerta=='CONSEJO DE LA JUDICATURA':
        sheet['D8'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.alerta=='DEFENSORIA':
        sheet['E8'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.alerta=='FISCALIA':
        sheet['F8'].fill = PatternFill(fill_type="solid", fgColor='009ee5')



    if devolucion_id.calidad_venta=='MALA VENTA':
        sheet['B9'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.calidad_venta=='NO ESTA DE ACUERDO CON EL PROCESO':
        sheet['C9'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.calidad_venta=='MUERTE':
        sheet['D9'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.calidad_venta=='ENFERMEDAD':
        sheet['E9'].fill = PatternFill(fill_type="solid", fgColor='009ee5')

    sheet = workbook['GERENCIA']
    listaSheet2g = list(filter(lambda x: (x['hoja']==5), lista)) 
    for campo in listaSheet2g:
       # fila=capturar_fila_de_valor_a_buscar_en_hoja_calculo(sheet,5,8,3,campo['valor'])
        cell = sheet.cell(row=campo['fila'], column=campo['columna'])
        try:
            cell.value = campo['valor'] or ''
        except:
            raise ValidationError("""El valor {0} en la fila {1} columna {2} hoja {3} se encuentra mal configurado en la plantilla""".format( campo['valor'], campo['fila'],campo['columna'],sheet))


    if devolucion_id.alerta=='CLIENTE':
        sheet['B8'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.alerta=='ABOGADO':
        sheet['C8'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.alerta=='CONSEJO DE LA JUDICATURA':
        sheet['D8'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.alerta=='DEFENSORIA':
        sheet['E8'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.alerta=='FISCALIA':
        sheet['F8'].fill = PatternFill(fill_type="solid", fgColor='009ee5')



    if devolucion_id.calidad_venta=='MALA VENTA':
        sheet['B9'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.calidad_venta=='NO ESTA DE ACUERDO CON EL PROCESO':
        sheet['C9'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.calidad_venta=='MUERTE':
        sheet['D9'].fill = PatternFill(fill_type="solid", fgColor='009ee5')
    elif devolucion_id.calidad_venta=='ENFERMEDAD':
        sheet['E9'].fill = PatternFill(fill_type="solid", fgColor='009ee5')

    workbook.save(ruta)




def llenar_tabla_excel(lista,sheet,filaInicioBusqueda,filafinBusqueda,ColumnaEje):

    for linea in lista:
        nombre_patrimonio=linea['nombre']


        fila=capturar_fila_de_valor_a_buscar_en_hoja_calculo(sheet,filaInicioBusqueda,filafinBusqueda,ColumnaEje,nombre_patrimonio)
        if fila:
            for campo in linea['campos']:
                cell = sheet.cell(row=fila, column=campo['columna'])
                print(fila,campo['columna'])
                cell.value = campo['valor']




def capturar_fila_de_valor_a_buscar_en_hoja_calculo(sheet,fila_ini,fila_fin,columna_eje,nombre_buscar):


    for fila in range(fila_ini,fila_fin+1):
        valor = sheet.cell(row=fila, column=columna_eje)

        print(valor,nombre_buscar)
        if valor.value==nombre_buscar:
            return fila
    return False



def Todalatabla(sheet, col, colfin, fil, filfin, styleleft, styletop, styleright, stylebottom):

    colfin=colfin+1
    filfin=filfin+2

    border_cell = Border(left=Side(style=styleleft), top=Side(style=styletop), right=Side(style=styleright), bottom=Side(style=stylebottom))
    for i in range(fil, filfin-1):
        for j in range(col, colfin):
            sheet.cell(row=i, column=j).border = border_cell




def border_cell(sheet, fil, col, styleleft, styletop, styleright, stylebottom):
    border_cell = Border(left=Side(style=styleleft), top=Side(style=styletop), right=Side(style=styleright), bottom=Side(style=stylebottom))
    sheet.cell(row=fil, column=col).border = border_cell
